from enum import Enum
import numpy as np
from scipy.optimize import nnls
from openpyxl import load_workbook

Kw = 1e-14
F = 96485
LOG10 = np.log(10)
TOLERANCE = 1e-6


class TitrationModes(Enum):
    VOLUMETRIC = 1
    COULOMETRIC = 2


class pKSpectrum:
    def __init__(self,
                 source_file,
                 mode: TitrationModes = TitrationModes.VOLUMETRIC):
        self.source_file = source_file
        self.mode = mode
        self.sample_name = None
        self.comment = None
        self.date = None
        self.time = None
        self.sample_volume = None
        self.alkaline_concentration = None
        self.current = None
        self.alkaline_volumes = []
        self.times = []
        self.ph_values = []
        self.alpha_values = []
        self.valid_points = 0
        self.acid_peaks = []

        self._load_data()

    def _load_data(self):
        """
        Loads data and makes a simple check
        :return: None
        """

        # Load workbook
        wb = load_workbook(self.source_file)
        ws = wb.active

        # Get sample information
        self.sample_name = ws['A1'].value
        self.comment = ws['A2'].value
        self.timestamp = ws['A3'].value
        self.sample_volume = ws['A4'].value
        if self.mode == TitrationModes.VOLUMETRIC:
            self.alkaline_concentration = ws['A5'].value
        else:
            self.current = ws['A5'].value

        # Get titration data
        shift = 0
        while True:
            volume = ws[f'A{6 + shift}'].value
            ph = ws[f'B{6 + shift}'].value

            if self._check_number(volume) and self._check_number(ph):
                self.alkaline_volumes.append(volume)
                self.ph_values.append(ph)
                shift += 1
            else:
                break

        # Arrange titration data
        swapped = False
        while True:
            for i in range(len(self.alkaline_volumes)-1):
                if self.alkaline_volumes[i] > self.alkaline_volumes[i+1]:
                    swapped = True
                    self.alkaline_volumes[i], self.alkaline_volumes[i+1] = \
                        self.alkaline_volumes[i+1], self.alkaline_volumes[i]
                    self.ph_values[i], self.ph_values[i+1] = self.ph_values[i+1], self.ph_values[i]
            if not swapped:
                break

        # Transform volume data to time if needed
        if self.mode == TitrationModes.COULOMETRIC:
            self.times = list(self.alkaline_volumes)
            self.alkaline_volumes = []

        # Check data validity
        for i in range(len(self.alkaline_volumes)):
            h = pow(10, -self.ph_values[i])
            if self.mode == TitrationModes.VOLUMETRIC:
                t = ((h - Kw / h) / self.sample_volume) * (self.alkaline_volumes[i] + self.sample_volume) + \
                    self.alkaline_concentration * self.alkaline_volumes[i] / self.sample_volume
            else:
                t = h - Kw / h + self.current * self.times[i] / F / self.sample_volume
            if t >= 0:
                self.alpha_values.append(t)
                self.valid_points = i + 1
            else:
                break

    def make_calculation(self, pk_start=0, pk_end=10, d_pk=0.05, integration_constant=True):
        """
        Calculates pK spectrum
        :param pk_start: Start pK value (float)
        :param pk_end: End pK value (float)
        :param d_pk: Delta pK value (float)
        :param integration_constant: Use integration constant? (boolean)
        :return: Peaks, calculation error
        """

        # Check for the valid points
        if self.valid_points < 7:
            return None, np.nan

        # Calculate constant step
        pk_step = round((pk_end - pk_start) / d_pk) + 1

        # Fill right part
        if integration_constant:
            shape_1 = pk_step + 2
        else:
            shape_1 = pk_step
        right = np.zeros((self.valid_points, shape_1))
        for i in range(self.valid_points):
            for j in range(pk_step):
                right[i, j] = d_pk / (1 + np.exp(LOG10 * (pk_start + d_pk * j - self.ph_values[i])))

        # Add items for the constant calculation
        if integration_constant:
            right[:, -2] = 1
            right[:, -1] = -1

        # Solve equation
        constants, residual = nnls(right, np.array(self.alpha_values))

        # Remove constant from scope
        if integration_constant:
            constants = constants[:-2]

        # Normalization
        constants *= d_pk

        # Truncate border artefacts
        if constants[0] > TOLERANCE > constants[1]:
            constants[0] = 0
        if constants[-1] > TOLERANCE > constants[-2]:
            constants[-1] = 0

        sum_constants = constants.sum()
        max_constant = constants.max(initial=0)
        threshold = max_constant / 100
        constants_relative = constants / sum_constants

        # Peak calculation sequence
        i = 0
        while i < pk_step:
            if constants[i] > threshold:
                self.acid_peaks.append({'point_count': 0, 'concentration': 0, 'first_point': i})
                while i < pk_step and constants[i] > threshold:
                    self.acid_peaks[-1]['point_count'] += 1
                    self.acid_peaks[-1]['concentration'] += constants[i]
                    i += 1
            else:
                i += 1

        # Peaks exact position and height calculation
        if len(self.acid_peaks) > 0:
            for i in range(len(self.acid_peaks)):
                t1 = 0
                t2 = 0
                peak = self.acid_peaks[i]
                for j in range(peak['point_count']):
                    t1 += constants_relative[peak['first_point'] + j] * \
                        (pk_start + d_pk * (peak['first_point'] + j))
                    t2 += constants_relative[peak['first_point'] + j]
                peak['mean'] = t1 / t2
            for i in range(len(self.acid_peaks)):
                peak = self.acid_peaks[i]
                if peak['point_count'] > 0:
                    t1 = 0
                    t2 = 0
                    for j in range(peak['point_count']):
                        t1 += constants_relative[peak['first_point'] + j] * \
                              (pk_start + d_pk * (peak['first_point'] + j) - peak['mean']) ** 2
                        t2 += constants_relative[peak['first_point'] + j]
                    peak['interval'] = 1.96 * np.sqrt(t1 / t2) / np.sqrt(peak['point_count'])
                else:
                    peak['interval'] = 0.

        # Calculate error
        error = np.sqrt(residual) / np.sqrt(pk_step - 1)

        return self.acid_peaks, error

    @staticmethod
    def _check_number(a):
        """
        Checks if argument is number
        :param a: Value to check (any)
        :return: Check result (boolean)
        """
        return type(a) == int or type(a) == float
